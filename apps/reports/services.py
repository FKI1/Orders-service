import csv
import io
import json
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum, Count, Avg, Q, F, Max, Min
from django.utils import timezone
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas

from .models import Report, ReportSchedule
from apps.orders.models import Order, OrderItem, Payment
from apps.products.models import Product, Category
from apps.users.models import User, UserActivity


class ReportGenerator:
    """
    Базовый класс для генерации отчетов.
    """
    
    @staticmethod
    def generate_report(report_id):
        """
        Генерация отчета по ID.
        """
        try:
            report = Report.objects.get(id=report_id)
            report.status = Report.ReportStatus.PROCESSING
            report.started_at = timezone.now()
            report.save()
            
            # Выбор метода генерации в зависимости от типа отчета
            if report.report_type.startswith('orders'):
                data = ReportGenerator.generate_orders_report(report)
            elif report.report_type.startswith('products'):
                data = ReportGenerator.generate_products_report(report)
            elif report.report_type.startswith('users'):
                data = ReportGenerator.generate_users_report(report)
            elif report.report_type.startswith('suppliers'):
                data = ReportGenerator.generate_suppliers_report(report)
            elif report.report_type.startswith('financial'):
                data = ReportGenerator.generate_financial_report(report)
            else:
                raise ValueError(f'Неизвестный тип отчета: {report.report_type}')
            
            # Генерация файла в нужном формате
            if report.format == Report.ReportFormat.EXCEL:
                file_content, filename = ReportGenerator.generate_excel(report, data)
            elif report.format == Report.ReportFormat.PDF:
                file_content, filename = ReportGenerator.generate_pdf(report, data)
            elif report.format == Report.ReportFormat.CSV:
                file_content, filename = ReportGenerator.generate_csv(report, data)
            elif report.format == Report.ReportFormat.JSON:
                file_content, filename = ReportGenerator.generate_json(report, data)
            else:
                raise ValueError(f'Неизвестный формат: {report.format}')
            
            # Сохраняем файл
            from django.core.files.base import ContentFile
            report.file.save(filename, ContentFile(file_content))
            report.file_size = report.file.size
            report.status = Report.ReportStatus.COMPLETED
            report.completed_at = timezone.now()
            report.save()
            
            # Отправляем уведомление
            ReportGenerator.send_notification(report)
            
            return report
            
        except Exception as e:
            report.status = Report.ReportStatus.FAILED
            report.error_message = str(e)
            report.completed_at = timezone.now()
            report.save()
            raise e
    
    @staticmethod
    def generate_orders_report(report):
        """
        Генерация отчета по заказам.
        """
        params = report.parameters
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        store_id = params.get('store_id')
        status = params.get('status')
        
        # Базовый запрос
        orders = Order.objects.all()
        
        if date_from:
            orders = orders.filter(created_at__date__gte=date_from)
        if date_to:
            orders = orders.filter(created_at__date__lte=date_to)
        if store_id:
            orders = orders.filter(store_id=store_id)
        if status:
            orders = orders.filter(status=status)
        
        # Общая статистика
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
        
        # Статусы заказов
        orders_by_status = orders.values('status').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        ).order_by('-count')
        
        # Заказы по дням
        orders_by_day = orders.extra(
            {'day': "date(created_at)"}
        ).values('day').annotate(
            count=Count('id'),
            revenue=Sum('total_amount')
        ).order_by('day')
        
        # Топ товаров
        top_products = OrderItem.objects.filter(
            order__in=orders
        ).values(
            'product__id', 'product__name', 'product__sku'
        ).annotate(
            quantity=Sum('quantity'),
            revenue=Sum('total')
        ).order_by('-quantity')[:10]
        
        return {
            'report_type': 'orders',
            'period': {
                'from': date_from,
                'to': date_to,
            },
            'summary': {
                'total_orders': total_orders,
                'total_revenue': float(total_revenue),
                'avg_order_value': float(avg_order_value),
                'total_items': orders.aggregate(total=Sum('items__quantity'))['total'] or 0,
            },
            'by_status': list(orders_by_status),
            'by_day': list(orders_by_day),
            'top_products': list(top_products),
            'orders': orders[:100],  # Ограничиваем для производительности
        }
    
    @staticmethod
    def generate_products_report(report):
        """
        Генерация отчета по товарам.
        """
        params = report.parameters
        report_type = report.report_type
        
        if report_type == Report.ReportType.PRODUCTS_INVENTORY:
            return ReportGenerator.generate_inventory_report(params)
        elif report_type == Report.ReportType.PRODUCTS_SALES:
            return ReportGenerator.generate_products_sales_report(params)
        elif report_type == Report.ReportType.PRODUCTS_POPULAR:
            return ReportGenerator.generate_popular_products_report(params)
        elif report_type == Report.ReportType.PRODUCTS_LOW_STOCK:
            return ReportGenerator.generate_low_stock_report(params)
    
    @staticmethod
    def generate_inventory_report(params):
        """
        Отчет по остаткам товаров.
        """
        category_id = params.get('category_id')
        supplier_id = params.get('supplier_id')
        
        products = Product.objects.filter(status=Product.ProductStatus.ACTIVE)
        
        if category_id:
            products = products.filter(category_id=category_id)
        if supplier_id:
            products = products.filter(supplier_id=supplier_id)
        
        # Общая статистика
        total_products = products.count()
        total_stock = products.aggregate(total=Sum('stock_quantity'))['total'] or 0
        total_value = products.aggregate(
            total=Sum(F('stock_quantity') * F('price'))
        )['total'] or 0
        
        # Товары с низким остатком
        low_stock = products.filter(
            stock_quantity__lte=F('min_stock_level')
        ).count()
        
        # Товары без остатка
        out_of_stock = products.filter(stock_quantity=0).count()
        
        # Распределение по категориям
        by_category = products.values(
            'category__id', 'category__name'
        ).annotate(
            count=Count('id'),
            stock=Sum('stock_quantity'),
            value=Sum(F('stock_quantity') * F('price'))
        ).order_by('-value')
        
        return {
            'report_type': 'inventory',
            'summary': {
                'total_products': total_products,
                'total_stock': total_stock,
                'total_value': float(total_value),
                'low_stock': low_stock,
                'out_of_stock': out_of_stock,
            },
            'by_category': list(by_category),
            'products': products.order_by('name')[:100],
        }
    
    @staticmethod
    def generate_users_report(report):
        """
        Генерация отчета по пользователям.
        """
        params = report.parameters
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        role = params.get('role')
        
        users = User.objects.filter(is_active=True)
        
        if date_from:
            users = users.filter(date_joined__date__gte=date_from)
        if date_to:
            users = users.filter(date_joined__date__lte=date_to)
        if role:
            users = users.filter(role=role)
        
        # Общая статистика
        total_users = users.count()
        
        # Регистрации по дням
        registrations_by_day = users.extra(
            {'day': "date(date_joined)"}
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day')
        
        # Распределение по ролям
        by_role = users.values('role').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Активные пользователи
        active_threshold = timezone.now() - timedelta(days=30)
        active_users = users.filter(last_login__gte=active_threshold).count()
        
        return {
            'report_type': 'users',
            'summary': {
                'total_users': total_users,
                'active_users': active_users,
                'inactive_users': total_users - active_users,
                'verified_users': users.filter(is_verified=True).count(),
                'email_verified': users.filter(email_verified=True).count(),
            },
            'by_role': list(by_role),
            'by_day': list(registrations_by_day),
            'users': users.order_by('-date_joined')[:100],
        }
    
    @staticmethod
    def generate_financial_report(report):
        """
        Генерация финансового отчета.
        """
        params = report.parameters
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        
        # Заказы за период
        orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            status=Order.Status.DELIVERED
        )
        
        # Платежи за период
        payments = Payment.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            status=Payment.PaymentStatus.COMPLETED
        )
        
        # Общая выручка
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        total_paid = payments.aggregate(total=Sum('amount'))['total'] or 0
        total_commission = payments.aggregate(total=Sum('commission'))['total'] or 0
        
        # Выручка по дням
        revenue_by_day = orders.extra(
            {'day': "date(created_at)"}
        ).values('day').annotate(
            revenue=Sum('total_amount'),
            orders=Count('id')
        ).order_by('day')
        
        # Платежи по методам
        payments_by_method = payments.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('amount'),
            commission=Sum('commission')
        ).order_by('-total')
        
        return {
            'report_type': 'financial',
            'period': {
                'from': date_from,
                'to': date_to,
            },
            'summary': {
                'total_revenue': float(total_revenue),
                'total_paid': float(total_paid),
                'total_commission': float(total_commission),
                'net_revenue': float(total_revenue - total_commission),
                'orders_count': orders.count(),
                'payments_count': payments.count(),
            },
            'by_day': list(revenue_by_day),
            'by_payment_method': list(payments_by_method),
        }
    
    @staticmethod
    def generate_excel(report, data):
        """
        Генерация Excel файла.
        """
        wb = Workbook()
        
        # Основной лист
        ws = wb.active
        ws.title = f"Отчет {report.report_number}"
        
        # Заголовок
        title_font = Font(size=14, bold=True)
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font_color = colors.white
        
        # Заголовок отчета
        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = f"{report.get_report_type_display()} - {report.created_at.strftime('%d.%m.%Y')}"
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Период
        if 'period' in data:
            ws.merge_cells('A2:E2')
            period = data['period']
            ws['A2'].value = f"Период: {period.get('from', 'начало')} - {period.get('to', 'конец')}"
            ws['A2'].font = Font(size=12, italic=True)
        
        row = 4
        
        # Сводка
        if 'summary' in data:
            ws.cell(row=row, column=1, value="СВОДКА").font = Font(size=12, bold=True)
            row += 1
            
            for key, value in data['summary'].items():
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 2
        
        # Таблица с данными
        if 'orders' in data:
            headers = ['№', 'Номер заказа', 'Дата', 'Сумма', 'Статус']
            ws.cell(row=row, column=1, value="ЗАКАЗЫ").font = Font(size=12, bold=True)
            row += 1
            
            # Заголовки таблицы
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row += 1
            
            # Данные
            for idx, order in enumerate(data['orders'], 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=order.order_number)
                ws.cell(row=row, column=3, value=order.created_at.strftime('%d.%m.%Y'))
                ws.cell(row=row, column=4, value=float(order.total_amount))
                ws.cell(row=row, column=5, value=order.get_status_display())
                row += 1
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"report_{report.report_number}.xlsx"
        return output.getvalue(), filename
    
    @staticmethod
    def generate_csv(report, data):
        """
        Генерация CSV файла.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Заголовок
        writer.writerow([f"{report.get_report_type_display()}"])
        writer.writerow([f"Дата: {report.created_at.strftime('%d.%m.%Y %H:%M')}"])
        
        if 'period' in data:
            period = data['period']
            writer.writerow([f"Период: {period.get('from', 'начало')} - {period.get('to', 'конец')}"])
        
        writer.writerow([])
        
        # Сводка
        if 'summary' in data:
            writer.writerow(['СВОДКА'])
            for key, value in data['summary'].items():
                writer.writerow([key.replace('_', ' ').title(), value])
            writer.writerow([])
        
        # Данные
        if 'orders' in data:
            writer.writerow(['№', 'Номер заказа', 'Дата', 'Сумма', 'Статус'])
            for idx, order in enumerate(data['orders'], 1):
                writer.writerow([
                    idx,
                    order.order_number,
                    order.created_at.strftime('%d.%m.%Y'),
                    float(order.total_amount),
                    order.get_status_display()
                ])
        
        output.seek(0)
        filename = f"report_{report.report_number}.csv"
        return output.getvalue().encode('utf-8'), filename
    
    @staticmethod
    def generate_pdf(report, data):
        """
        Генерация PDF файла.
        """
        output = io.BytesIO()
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,  # Center
            spaceAfter=20
        )
        
        elements.append(Paragraph(
            f"{report.get_report_type_display()}",
            title_style
        ))
        
        # Дата
        date_style = ParagraphStyle(
            'Date',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=20
        )
        
        elements.append(Paragraph(
            f"Сгенерирован: {report.created_at.strftime('%d.%m.%Y %H:%M')}",
            date_style
        ))
        
        # Период
        if 'period' in data:
            period = data['period']
            elements.append(Paragraph(
                f"Период: {period.get('from', 'начало')} - {period.get('to', 'конец')}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.5*cm))
        
        # Сводка
        if 'summary' in data:
            elements.append(Paragraph("Сводка", styles['Heading2']))
            elements.append(Spacer(1, 0.3*cm))
            
            summary_data = []
            for key, value in data['summary'].items():
                summary_data.append([
                    key.replace('_', ' ').title(),
                    str(value)
                ])
            
            summary_table = Table(summary_data, colWidths=[6*cm, 6*cm])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*cm))
        
        # Сборка PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"report_{report.report_number}.pdf"
        return output.getvalue(), filename
    
    @staticmethod
    def generate_json(report, data):
        """
        Генерация JSON файла.
        """
        output = {
            'report_number': report.report_number,
            'report_type': report.report_type,
            'report_type_display': report.get_report_type_display(),
            'created_at': report.created_at.isoformat(),
            'parameters': report.parameters,
            'data': data
        }
        
        json_data = json.dumps(output, indent=2, ensure_ascii=False, default=str)
        filename = f"report_{report.report_number}.json"
        return json_data.encode('utf-8'), filename
    
    @staticmethod
    def send_notification(report):
        """
        Отправка уведомления о готовности отчета.
        """
        if not report.created_by or not report.created_by.email:
            return
        
        subject = f"Отчет готов: {report.get_report_type_display()}"
        
        context = {
            'report': report,
            'user': report.created_by,
        }
        
        html_message = render_to_string('emails/report_ready.html', context)
        plain_message = render_to_string('emails/report_ready.txt', context)
        
        email = EmailMessage(
            subject=subject,
            body=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[report.created_by.email],
        )
        
        if report.file:
            email.attach_file(report.file.path)
        
        email.send()


class ReportScheduler:
    """
    Планировщик автоматической генерации отчетов.
    """
    
    @staticmethod
    def run_scheduled_reports():
        """
        Запуск отчетов по расписанию.
        """
        now = timezone.now()
        schedules = ReportSchedule.objects.filter(
            is_active=True,
            next_run__lte=now
        )
        
        for schedule in schedules:
            try:
                # Создаем отчет
                report = Report.objects.create(
                    report_type=schedule.report_type,
                    format=schedule.format,
                    parameters=schedule.parameters,
                    created_by=schedule.created_by,
                )
                
                # Запускаем генерацию асинхронно
                from config.celery import current_app
                current_app.send_task('generate_report', args=[report.id])
                
                # Обновляем расписание
                schedule.last_run = now
                schedule.next_run = ReportScheduler.calculate_next_run(schedule)
                schedule.save()
                
            except Exception as e:
                print(f"Error running scheduled report {schedule.id}: {e}")
    
    @staticmethod
    def calculate_next_run(schedule):
        """
        Расчет следующего запуска.
        """
        now = timezone.now()
        
        if schedule.frequency == ReportSchedule.Frequency.DAILY:
            return now + timedelta(days=1)
        elif schedule.frequency == ReportSchedule.Frequency.WEEKLY:
            return now + timedelta(weeks=1)
        elif schedule.frequency == ReportSchedule.Frequency.MONTHLY:
            # Первый день следующего месяца
            if now.month == 12:
                return now.replace(year=now.year + 1, month=1, day=1)
            else:
                return now.replace(month=now.month + 1, day=1)
        elif schedule.frequency == ReportSchedule.Frequency.QUARTERLY:
            # Первый день следующего квартала
            quarter = (now.month - 1) // 3 + 1
            next_quarter = quarter + 1
            if next_quarter > 4:
                return now.replace(year=now.year + 1, month=1, day=1)
            else:
                return now.replace(month=(next_quarter - 1) * 3 + 1, day=1)
        elif schedule.frequency == ReportSchedule.Frequency.YEARLY:
            return now.replace(year=now.year + 1, month=1, day=1)
        
        return now + timedelta(days=1)
